#!/usr/bin/env python3
"""
SmartKinneret preprocessing (authoritative data generator).

Reads the raw monitoring files in data_raw/, normalizes them into a common
schema, applies the agreed unit/quality verification rules, aggregates by
hour/day/month, computes the rule-based risk model and the statistical tests,
and writes normalized JSON to public/data/processed/.

A Node equivalent (scripts/preprocess.mjs) reproduces this for `npm run preprocess`;
this Python version is what ships the bundled processed data.

Data-handling contract (confirmed with project owner):
  * Bteha / Zemah wind speed treated as m/s -- DOCUMENTED ASSUMPTION.
  * Ein Gev air temperature: NOT converted, tagged 'unit/quality unverified',
    excluded from stats / risk / ML.
  * Ginosar wind speed: tagged 'unit/quality unverified', excluded likewise.
  * Stations kept; only verified variables used/displayed.
  * Depth(mm) is a sensor measurement field, never Lake Kinneret water level.
"""
import json, math, os, sys, datetime as dt
from pathlib import Path
import numpy as np
import pandas as pd
from scipy import stats as sps
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data_raw"
OUT = ROOT / "public" / "data" / "processed"
OUT.mkdir(parents=True, exist_ok=True)

RISK_VERSION = "1.0.0"
NA_TOKENS = {"-", "", "NAN", "nan", "NaN", "null", "NA", "N/A"}
parse_errors = []   # collected parse-level warnings -> error report

def num(series):
    """Coerce to numeric, treating documented NA tokens as missing."""
    s = series.astype(str).str.strip().replace(list(NA_TOKENS), np.nan)
    return pd.to_numeric(s, errors="coerce")

def jround(x, n=3):
    if x is None: return None
    if isinstance(x, (float, np.floating)):
        if math.isnan(x) or math.isinf(x): return None
        return round(float(x), n)
    if isinstance(x, (int, np.integer)): return int(x)
    return x

def iso(ts):
    if pd.isna(ts): return None
    return pd.Timestamp(ts).strftime("%Y-%m-%dT%H:%M:%S")

# ---------------------------------------------------------------------------
# Station metadata. Marker coordinates are APPROXIMATE for the stylized map
# only -- NOT survey-grade GPS. (lat/lng used to place SVG markers.)
# ---------------------------------------------------------------------------
STATIONS = {
    "bteha":   {"name": "Bteha",   "type": "meteorological", "lat": 32.876, "lng": 35.638, "approxCoords": True},
    "zemah":   {"name": "Zemah",   "type": "meteorological", "lat": 32.704, "lng": 35.585, "approxCoords": True},
    "ginosar": {"name": "Ginosar", "type": "meteorological", "lat": 32.844, "lng": 35.512, "approxCoords": True},
    "eingev":  {"name": "Ein Gev", "type": "meteorological", "lat": 32.779, "lng": 35.646, "approxCoords": True},
    "knw":     {"name": "Golan Beach (KNW)", "type": "wave+current", "lat": 32.842, "lng": 35.651, "approxCoords": True},
    "knc":     {"name": "Station F (KNC)",   "type": "current",      "lat": 32.810, "lng": 35.595, "approxCoords": True},
    "metr_a":    {"name": "Metr-A (2025)",    "type": "meteorological+lake", "lat": 32.820, "lng": 35.600, "approxCoords": True},
    "a_probe":   {"name": "A-Probe (2025)",  "type": "water-quality",       "lat": 32.820, "lng": 35.600, "approxCoords": True},
    "kfar_nahum":{"name": "Kfar Nahum",      "type": "meteorological",      "lat": 32.869, "lng": 35.555, "approxCoords": True},
    "metr":      {"name": "Metr (2025)",      "type": "meteorological+lake", "lat": 32.821, "lng": 35.601, "approxCoords": True},
}

# variable -> display unit (verified)
UNITS = {
    "windSpeed": "m/s", "windDir": "deg", "airTemp": "C", "humidity": "%",
    "rainfall": "mm", "waveHeight": "m", "wavePeriod": "s", "waveDir": "deg",
    "currentMag": "cm/s", "currentDir": "deg", "sensorDepth": "mm",
    "waterTemp": "C", "lightLevel": "lux", "pressure": "mbar",
    "spCond": "µS/cm", "dissolvedO2": "mg/L", "turbidity": "FNU",
    "chlorophyll": "µg/L", "orp": "mV",
    "waveHeightMax": "m", "waveHeightTop10": "m", "wavePeriodMean": "s",
}

# Columns we deliberately exclude (unit/quality unverified).
EXCLUSIONS = []   # filled during parsing

def flag(station, column, variable, reason, action):
    EXCLUSIONS.append({"station": station, "column": column, "variable": variable,
                       "reason": reason, "action": action})

# ===========================================================================
# 1. METEO PARSERS  (column-name normalization differs per station)
# ===========================================================================
def parse_bteha():
    f = RAW / "Meteo" / "Meto_Bteha2024.csv"
    df = pd.read_csv(f)
    raw_rows = len(df)
    ts = pd.to_datetime(df["Date_Time"], format="%d/%m/%Y %H:%M", errors="coerce")
    out = pd.DataFrame({
        "timestamp": ts,
        "humidity": num(df["Rh"]),
        "airTemp": num(df["td_B"]),       # Celsius (plausible range) -> verified
        "windDir": num(df["Wd"]),
        "windSpeed": num(df["ws_B"]),     # assumed m/s (documented)
        "rainfall": num(df["Rain"]),
    })
    return finalize_meteo("bteha", out, raw_rows, f.name,
                          verified=["humidity","airTemp","windDir","windSpeed","rainfall"],
                          ws_assumed=True)

def parse_zemah():
    f = RAW / "Meteo" / "Meto_Zemah2024.csv"
    df = pd.read_csv(f)
    raw_rows = len(df)
    ts = pd.to_datetime(df["date"], format="%d/%m/%Y %H:%M", errors="coerce")
    out = pd.DataFrame({
        "timestamp": ts,
        "humidity": num(df["rh_Z"]),
        "airTemp": num(df["Ta_Z"]),
        "windDir": num(df["Wd_Z"]),
        "windSpeed": num(df["Ws_Z"]),     # assumed m/s (documented)
        "rainfall": num(df["Rain_Z"]),
    })
    return finalize_meteo("zemah", out, raw_rows, f.name,
                          verified=["humidity","airTemp","windDir","windSpeed","rainfall"],
                          ws_assumed=True)

def parse_ginosar():
    # legacy .xls was converted to xlsx with LibreOffice during inspection
    f = RAW / "Meteo" / "Meto_Ginosar2024_converted.xlsx"
    wb = load_workbook(f, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    raw_rows = len(rows)
    recs = []
    for r in rows:
        d, t = r[0], r[1]
        if d is None: continue
        # combine separate date + time columns
        try:
            base = pd.Timestamp(d).normalize()
            if t is not None:
                base = base + pd.Timedelta(hours=getattr(t, "hour", 0),
                                           minutes=getattr(t, "minute", 0),
                                           seconds=getattr(t, "second", 0))
        except Exception:
            base = pd.NaT
        recs.append((base, r[2], r[3], r[4], r[5]))
    out = pd.DataFrame(recs, columns=["timestamp","airTemp","humidity","windSpeed","windDir"])
    for c in ["airTemp","humidity","windSpeed","windDir"]:
        out[c] = num(out[c])
    # Ginosar wind speed is stuck near zero all year -> UNVERIFIED, excluded.
    flag("ginosar", "WS_ms_Avg", "windSpeed",
         "Wind speed reads ~0.000 m/s all year (median 0.000, max 0.94) while other "
         "stations average ~1.7 m/s; consistent with a frozen/failed anemometer.",
         "Excluded from statistics, risk and ML. Displayed only as a flagged quality issue.")
    out["windSpeed"] = np.nan   # drop unverified column from usable data
    return finalize_meteo("ginosar", out, raw_rows, "Meto_Ginosar2024.xls",
                          verified=["airTemp","humidity","windDir"],
                          unverified=["windSpeed"])

def parse_eingev():
    f = RAW / "Meteo" / "Ein_Gev2022-2024.dat"
    # TOA5: row0 metadata, row1 names, row2 units, row3 aggregation, then data
    df = pd.read_csv(f, skiprows=[0,2,3])
    raw_rows = len(df)
    ts = pd.to_datetime(df["TIMESTAMP"], errors="coerce")
    out = pd.DataFrame({
        "timestamp": ts,
        "airTemp": num(df["AirTC_Avg"]),    # UNVERIFIED unit -> excluded below
        "humidity": num(df["RH_Max"]),
        "windSpeed": num(df["WS_ms_Avg"]),  # documented m/s in file units row -> verified
        "windDir": num(df["WindDir"]),
        "rainfall": num(df["Rain_mm_Tot"]),
    })
    # Ein Gev air temperature inconsistent with Celsius (median ~81.7, range -241..156).
    flag("eingev", "AirTC_Avg", "airTemp",
         "Stored air-temperature values are physically impossible as Celsius (median 81.7, "
         "essentially all values >50 C, range -241..156). Possibly Fahrenheit or a "
         "miscalibrated channel; not confirmed by file metadata.",
         "NOT converted. Excluded from statistics, risk and ML. Displayed only as a flagged quality issue.")
    out["airTemp"] = np.nan   # drop unverified column from usable data
    return finalize_meteo("eingev", out, raw_rows, f.name,
                          verified=["humidity","windSpeed","windDir","rainfall"],
                          unverified=["airTemp"])

def parse_a_probe():
    f = RAW / "Meteo" / "Daily_Summary_a_probe_2025.csv"
    df = pd.read_csv(f)
    raw_rows = len(df)
    ts = pd.to_datetime(df["Date"], format="%Y-%m-%d", errors="coerce") + pd.Timedelta(hours=12)
    out = pd.DataFrame({
        "timestamp":   ts,
        "waterTemp":   num(df["Avg_Temp_(C)"]),
        "spCond":      num(df["Avg_SpCond_(us/cm)"]),
        "dissolvedO2": num(df["Avg_DO_(mg/L)"]),
        "turbidity":   num(df["Avg_Turbidity_(FNU)"]),
        "chlorophyll": num(df["Avg_Chlorophyll_(ug/L)"]),
        "orp":         num(df["Avg_ORP_(mV)"]),
    })
    return finalize_meteo("a_probe", out, raw_rows, f.name,
                          verified=["waterTemp","spCond","dissolvedO2","turbidity","chlorophyll","orp"],
                          daily_summary=True)

def parse_knw_daily_waves():
    files_info = [
        (RAW / "Lake_waves_currents" / "Daily_Summary_KNW08_Waves.csv", "KNW08"),
        (RAW / "Lake_waves_currents" / "Daily_Summary_KNW09_Waves.csv", "KNW09"),
        (RAW / "Lake_waves_currents" / "Daily_Summary_KNW10_Waves.csv", "KNW10"),
        (RAW / "Lake_waves_currents" / "Daily_Summary_KNW11_Waves.csv", "KNW11"),
    ]
    frames, raw_total, per_file = [], 0, []
    for f, label in files_info:
        df = pd.read_csv(f)
        raw_total += len(df)
        ts = pd.to_datetime(df["Date"], format="%Y-%m-%d", errors="coerce") + pd.Timedelta(hours=12)
        sub = pd.DataFrame({
            "timestamp":      ts,
            "waveHeight":     num(df["Avg_Sig_Wave_Height_(m)"]),
            "waveHeightMax":  num(df["Max_Sig_Wave_Height_(m)"]),
            "waveHeightTop10":num(df["Max_Top_10%_Wave_Height_(m)"]),
            "wavePeriod":     num(df["Avg_Peak_Period_(sec)"]),
            "wavePeriodMean": num(df["Avg_Mean_Period_(sec)"]),
            "sensorDepth":    num(df["Avg_Depth_(mm)"]),
        })
        sub = sub.dropna(subset=["timestamp"])
        per_file.append({"file": f.name, "rows": len(df),
                         "range": [iso(sub["timestamp"].min()), iso(sub["timestamp"].max())]})
        frames.append(sub)
    allw = pd.concat(frames, ignore_index=True).sort_values("timestamp").reset_index(drop=True)
    # On overlap days (KNW08/KNW09 transition) keep KNW08 entry (first in sort order)
    allw = allw.drop_duplicates(subset=["timestamp"], keep="first")
    return {"df": allw, "raw_rows": raw_total, "per_file": per_file}

def parse_metr_a():
    f = RAW / "Meteo" / "Daily_Summary_metr_a_2025.csv"
    df = pd.read_csv(f)
    raw_rows = len(df)
    # Daily summary data — timestamp set to noon to represent the daily average
    ts = pd.to_datetime(df["Date"], format="%Y-%m-%d", errors="coerce") + pd.Timedelta(hours=12)
    out = pd.DataFrame({
        "timestamp":  ts,
        "waterTemp":  num(df["Avg_Water_Temp_(C)"]),
        "airTemp":    num(df["Avg_Air_Temp_(C)"]),
        "windSpeed":  num(df["Avg_Wind_Speed_(m/s)"]),
        "humidity":   num(df["Avg_Humidity_(%)"]),
        "pressure":   num(df["Avg_Pressure_(mbar)"]),
        "rainfall":   num(df["Total_Rainfall_(mm)"]),
        "lightLevel": num(df["Avg_Light_Level"]),
    })
    return finalize_meteo("metr_a", out, raw_rows, f.name,
                          verified=["waterTemp","airTemp","windSpeed","humidity","pressure","rainfall","lightLevel"],
                          daily_summary=True)

def parse_kfar_nahum():
    f = RAW / "Meteo" / "Daily_Summary_KfarNahum2025.csv"
    df = pd.read_csv(f)
    raw_rows = len(df)
    ts = pd.to_datetime(df["day"], format="%Y-%m-%d", errors="coerce") + pd.Timedelta(hours=12)
    out = pd.DataFrame({
        "timestamp": ts,
        "airTemp":   num(df["temp_avg"]),
        "humidity":  num(df["humidity_avg"]),
        "windSpeed": np.nan,   # only max wind speed available; excluded to avoid biasing risk model
        "rainfall":  num(df["total_rain"]),
    })
    return finalize_meteo("kfar_nahum", out, raw_rows, f.name,
                          verified=["airTemp","humidity","rainfall"],
                          daily_summary=True)

def parse_metr():
    f = RAW / "Meteo" / "daily_metr_summary_2025.csv"
    df = pd.read_csv(f)
    raw_rows = len(df)
    ts = pd.to_datetime(df["Date"], format="%Y-%m-%d", errors="coerce") + pd.Timedelta(hours=12)
    out = pd.DataFrame({
        "timestamp": ts,
        "waterTemp": num(df["Water_Temp_Avg_C"]),
        "windSpeed": num(df["Wind_Speed_Avg_ms"]),
        "windDir":   num(df["Wind_Direction_Avg_deg"]),
        "airTemp":   num(df["Air_Temp_Avg_C"]),
        "humidity":  num(df["Humidity_Avg_%"]),
        "rainfall":  num(df["Rainfall_Sum_mm"]),
        "pressure":  num(df["Pressure_Avg_mbar"]),
    })
    return finalize_meteo("metr", out, raw_rows, f.name,
                          verified=["waterTemp","windSpeed","windDir","airTemp","humidity","rainfall","pressure"],
                          daily_summary=True)

def finalize_meteo(station, df, raw_rows, filename, verified, unverified=None, ws_assumed=False, daily_summary=False):
    df = df.dropna(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)
    return {"station": station, "df": df, "raw_rows": raw_rows, "filename": filename,
            "verified": verified, "unverified": unverified or [], "ws_assumed": ws_assumed,
            "daily_summary": daily_summary, "kind": "meteo"}

def parse_bteha_daily():
    """Combined 2025+2026 daily Bteha summaries for extending the hourly bteha time series."""
    frames, raw_total = [], 0
    f25 = RAW / "Meteo" / "Daily_Summary_Bteha2025.csv"
    if f25.exists():
        df25 = pd.read_csv(f25)
        raw_total += len(df25)
        ts = pd.to_datetime(df25["Date"], format="%Y-%m-%d", errors="coerce") + pd.Timedelta(hours=12)
        frames.append(pd.DataFrame({
            "timestamp": ts,
            "airTemp":   num(df25["Avg_Temp"]),
            "humidity":  num(df25["Avg_Humidity"]),
            "windSpeed": num(df25["Avg_Wind_Speed"]),
            "rainfall":  num(df25["Total_Rainfall"]),
        }))
    f26 = RAW / "Meteo" / "Daily_Summary_Bteha2026.csv"
    if f26.exists():
        df26 = pd.read_csv(f26)
        raw_total += len(df26)
        ts = pd.to_datetime(df26["day"], format="%Y-%m-%d", errors="coerce") + pd.Timedelta(hours=12)
        frames.append(pd.DataFrame({
            "timestamp": ts,
            "airTemp":   num(df26["temp_avg"]),
            "humidity":  num(df26["humidity_avg"]),
            "windSpeed": np.nan,   # only max available in 2026 file; excluded to avoid biasing risk model
            "rainfall":  num(df26["total_rain"]),
        }))
    if not frames:
        return pd.DataFrame(), 0
    combined = pd.concat(frames, ignore_index=True).sort_values("timestamp").reset_index(drop=True)
    return combined.dropna(subset=["timestamp"]), raw_total

def parse_zemah_daily():
    """Combined 2025+2026 daily Zemah summaries for extending the hourly zemah time series."""
    frames, raw_total = [], 0
    for fname in ["daily_meto_zemah_summary_2025.csv", "daily_meto_zemah_summary_2026.csv"]:
        f = RAW / "Meteo" / fname
        if f.exists():
            df = pd.read_csv(f)
            raw_total += len(df)
            ts = pd.to_datetime(df["Date"], format="%Y-%m-%d", errors="coerce") + pd.Timedelta(hours=12)
            frames.append(pd.DataFrame({
                "timestamp": ts,
                "airTemp":   num(df["Air_Temp_Avg_C"]),
                "windSpeed": num(df["Wind_Speed_Avg_ms"]),
                "windDir":   num(df["Wind_Direction_Avg_deg"]),
                "humidity":  num(df["Humidity_Avg_%"]),
                "rainfall":  num(df["Rainfall_Sum_mm"]),
            }))
    if not frames:
        return pd.DataFrame(), 0
    combined = pd.concat(frames, ignore_index=True).sort_values("timestamp").reset_index(drop=True)
    return combined.dropna(subset=["timestamp"]), raw_total

# ===========================================================================
# 2. CURRENTS PARSER (depth profiles -> surface + depth-averaged per timestamp)
# ===========================================================================
def parse_currents_station(prefix):
    files = sorted((RAW / "Lake_waves_currents").glob(f"{prefix}*_Currents.csv"))
    frames, depth_levels, raw_total, per_file = [], set(), 0, []
    for f in files:
        df = pd.read_csv(f)
        raw_total += len(df)
        ts = pd.to_datetime(df["Date"], errors="coerce")
        depth = num(df["Depth"])
        mag = num(df["Magnitude [cm/s]"])
        direction = num(df["Direction [Degrees]"])
        sub = pd.DataFrame({"timestamp": ts, "depth": depth, "currentMag": mag, "currentDir": direction})
        sub = sub.dropna(subset=["timestamp"])
        depth_levels |= set(sub["depth"].dropna().round(2).unique().tolist())
        per_file.append({"file": f.name, "rows": len(df),
                         "range": [iso(sub["timestamp"].min()), iso(sub["timestamp"].max())]})
        frames.append(sub)
    allp = pd.concat(frames, ignore_index=True)
    # Surface = shallowest depth row per timestamp; depth-avg = mean across column.
    allp = allp.sort_values(["timestamp","depth"])
    surface = allp.groupby("timestamp", as_index=False).first()[["timestamp","currentMag","currentDir","depth"]]
    surface = surface.rename(columns={"currentMag":"currentMag","currentDir":"currentDir","depth":"surfaceDepth"})
    davg = allp.groupby("timestamp", as_index=False)["currentMag"].mean().rename(columns={"currentMag":"currentMagAvg"})
    merged = surface.merge(davg, on="timestamp", how="left").sort_values("timestamp").reset_index(drop=True)
    return {"df": merged, "depth_levels": sorted(depth_levels), "raw_rows": raw_total,
            "per_file": per_file}

# ===========================================================================
# 3. WAVES PARSER (KNW only) -> hourly Hs / Tp / Dir / sensor Depth(mm)
# ===========================================================================
def parse_waves_station(prefix):
    files = sorted((RAW / "Lake_waves_currents").glob(f"{prefix}*_Waves.xlsx"))
    frames, raw_total, per_file = [], 0, []
    for f in files:
        wb = load_workbook(f, read_only=True); ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True); hdr = list(next(it))
        idx = {name: i for i, name in enumerate(hdr)}
        def col(*names):
            for n in names:
                if n in idx: return idx[n]
            return None
        i_ts = col("Date & Time (GMT)"); i_hs = col("Hs (m)"); i_tp = col("Tp (sec)")
        i_dp = col("DirPeak(deg)"); i_dep = col("Depth(mm)")
        rows = []
        for r in it:
            if i_ts is None or r[i_ts] is None: continue
            rows.append((r[i_ts],
                         r[i_hs] if i_hs is not None else None,
                         r[i_tp] if i_tp is not None else None,
                         r[i_dp] if i_dp is not None else None,
                         r[i_dep] if i_dep is not None else None))
        sub = pd.DataFrame(rows, columns=["timestamp_gmt","waveHeight","wavePeriod","waveDir","sensorDepth"])
        raw_total += len(sub)
        per_file.append({"file": f.name, "rows": len(sub)})
        frames.append(sub)
    allw = pd.concat(frames, ignore_index=True)
    allw["timestamp_gmt"] = pd.to_datetime(allw["timestamp_gmt"], errors="coerce")
    # Waves are GMT/UTC. Convert to Israel local (UTC+2 standard / +3 DST) to align
    # with the unlabeled meteo/current series (assumed local). Documented conversion.
    g = allw["timestamp_gmt"].dt.tz_localize("UTC")
    allw["timestamp"] = g.dt.tz_convert("Asia/Jerusalem").dt.tz_localize(None)
    for c in ["waveHeight","wavePeriod","waveDir","sensorDepth"]:
        allw[c] = num(allw[c])
    allw = allw.dropna(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)
    return {"df": allw[["timestamp","waveHeight","wavePeriod","waveDir","sensorDepth"]],
            "raw_rows": raw_total, "per_file": per_file}

# ===========================================================================
# AGGREGATION helpers
# ===========================================================================
def aggregate(df, value_cols, freq):
    if df.empty: return {"timestamps": [], **{c: [] for c in value_cols}}
    g = df.set_index("timestamp").resample(freq)[value_cols].mean()
    g = g.dropna(how="all")
    out = {"timestamps": [iso(t) for t in g.index]}
    for c in value_cols:
        out[c] = [jround(v, 3) if pd.notna(v) else None for v in g[c].values]
    return out

def describe(series):
    s = pd.to_numeric(series, errors="coerce")
    valid = s.dropna()
    if len(valid) == 0:
        return {"count":0,"missing":int(s.isna().sum()),"mean":None,"median":None,
                "min":None,"max":None,"std":None,"p25":None,"p50":None,"p75":None}
    return {"count": int(len(valid)), "missing": int(s.isna().sum()),
            "mean": jround(valid.mean()), "median": jround(valid.median()),
            "min": jround(valid.min()), "max": jround(valid.max()),
            "std": jround(valid.std(ddof=1)) if len(valid) > 1 else None,
            "p25": jround(valid.quantile(.25)), "p50": jround(valid.quantile(.50)),
            "p75": jround(valid.quantile(.85))}

print("Parsing meteo stations ...")
meteo = {m["station"]: m for m in [parse_bteha(), parse_zemah(), parse_ginosar(), parse_eingev(), parse_metr_a(), parse_a_probe(), parse_kfar_nahum(), parse_metr()]}
print("Extending Bteha with 2025/2026 daily summaries ...")
_bteha_ext, _bteha_ext_rows = parse_bteha_daily()
if not _bteha_ext.empty:
    _bteha_cutoff = meteo["bteha"]["df"]["timestamp"].max()
    _new_bteha = _bteha_ext[_bteha_ext["timestamp"] > _bteha_cutoff].copy()
    if not _new_bteha.empty:
        for c in meteo["bteha"]["df"].columns:
            if c not in _new_bteha.columns:
                _new_bteha[c] = np.nan
        meteo["bteha"]["df"] = pd.concat(
            [meteo["bteha"]["df"], _new_bteha], ignore_index=True
        ).sort_values("timestamp").reset_index(drop=True)
        meteo["bteha"]["raw_rows"] += _bteha_ext_rows
        print(f"  Bteha extended by {len(_new_bteha)} daily rows to {meteo['bteha']['df']['timestamp'].max().date()}")
print("Extending Zemah with 2025/2026 daily summaries ...")
_zemah_ext, _zemah_ext_rows = parse_zemah_daily()
if not _zemah_ext.empty:
    _zemah_cutoff = meteo["zemah"]["df"]["timestamp"].max()
    _new_zemah = _zemah_ext[_zemah_ext["timestamp"] > _zemah_cutoff].copy()
    if not _new_zemah.empty:
        for c in meteo["zemah"]["df"].columns:
            if c not in _new_zemah.columns:
                _new_zemah[c] = np.nan
        meteo["zemah"]["df"] = pd.concat(
            [meteo["zemah"]["df"], _new_zemah], ignore_index=True
        ).sort_values("timestamp").reset_index(drop=True)
        meteo["zemah"]["raw_rows"] += _zemah_ext_rows
        print(f"  Zemah extended by {len(_new_zemah)} daily rows to {meteo['zemah']['df']['timestamp'].max().date()}")
print("Parsing lake currents ...")
knw_cur = parse_currents_station("KNW")
knc_cur = parse_currents_station("KNC")
print("Parsing lake waves ...")
knw_wav = parse_waves_station("KNW")
print("Parsing KNW daily wave summaries (2025) ...")
knw_wav_daily = parse_knw_daily_waves()
# Extend existing KNW wave data with new daily summaries (keep only strictly new dates)
_knw_end = knw_wav["df"]["timestamp"].max() if not knw_wav["df"].empty else pd.Timestamp("2024-12-30")
_new_waves = knw_wav_daily["df"][knw_wav_daily["df"]["timestamp"] > _knw_end].copy()
if not _new_waves.empty:
    _new_waves["waveDir"] = np.nan   # daily summaries have no directional data
    knw_wav["df"] = pd.concat([knw_wav["df"], _new_waves], ignore_index=True).sort_values("timestamp").reset_index(drop=True)
    knw_wav["raw_rows"] += knw_wav_daily["raw_rows"]
    knw_wav["per_file"].extend(knw_wav_daily["per_file"])

# Build unified KNW lake frame (hourly waves + hourly surface current), KNC (current only)
def hourly(df, cols):
    if df.empty: return df
    # Only resample columns that actually exist in df
    cols = [c for c in cols if c in df.columns]
    if not cols: return pd.DataFrame(columns=["timestamp"])
    return df.set_index("timestamp").resample("1h")[cols].mean().dropna(how="all").reset_index()

_wave_extra = [c for c in ["waveHeightMax","waveHeightTop10","wavePeriodMean"]
               if c in knw_wav["df"].columns and knw_wav["df"][c].notna().any()]
_wave_cols = ["waveHeight","wavePeriod","waveDir","sensorDepth"] + _wave_extra
knw_lake = pd.merge(
    hourly(knw_wav["df"], _wave_cols),
    hourly(knw_cur["df"], ["currentMag","currentMagAvg","currentDir"]),
    on="timestamp", how="outer").sort_values("timestamp").reset_index(drop=True)
knc_lake = hourly(knc_cur["df"], ["currentMag","currentMagAvg","currentDir"]).sort_values("timestamp").reset_index(drop=True)

# ===========================================================================
# RISK MODEL (data-driven percentile thresholds; weights from config mirror)
# Verified inputs only: windSpeed, waveHeight, currentMag(surface), rainfall.
# ===========================================================================
RISK_WEIGHTS = {"windSpeed": 30, "waveHeight": 35, "currentMag": 20, "rainfall": 15}

def pct_thresholds(values):
    v = pd.to_numeric(pd.Series(values), errors="coerce").dropna()
    if len(v) < 20: return None
    return {"p50": jround(v.quantile(.50)), "p85": jround(v.quantile(.85))}

# Reference distributions for thresholds (verified data only):
wind_ref = pd.concat([meteo["bteha"]["df"]["windSpeed"], meteo["zemah"]["df"]["windSpeed"],
                      meteo["eingev"]["df"]["windSpeed"]]).dropna()
THRESH = {
    "windSpeed": pct_thresholds(wind_ref),
    "waveHeight": pct_thresholds(knw_wav["df"]["waveHeight"]),
    "currentMag": pct_thresholds(pd.concat([knw_cur["df"]["currentMag"], knc_cur["df"]["currentMag"]]).dropna()),
    "rainfall": {"p50": 0.0, "p85": 0.2},   # rainfall is zero-inflated; fixed light/heavy cut
}

def var_points(var, value):
    """0..weight points: 0 below p50, linear p50->p85 to half, p85+ scales to full."""
    if value is None or (isinstance(value, float) and math.isnan(value)): return None
    t = THRESH.get(var)
    w = RISK_WEIGHTS[var]
    if not t: return None
    p50, p85 = t["p50"], t["p85"]
    if value <= p50: return 0.0
    if value <= p85:
        return round(w * 0.5 * (value - p50) / max(p85 - p50, 1e-9), 2)
    # above p85: scale remaining half up to ~2*p85, capped at full weight
    extra = min((value - p85) / max(p85 - p50, 1e-9), 1.0)
    return round(w * (0.5 + 0.5 * extra), 2)

def compute_risk(inputs):
    """inputs: dict var->value (verified). Returns score/category/contributions."""
    contribs, total, available = [], 0.0, 0
    for var, w in RISK_WEIGHTS.items():
        val = inputs.get(var)
        pts = var_points(var, val)
        if pts is None:
            continue
        available += 1
        total += pts
        if pts > 0:
            label = {"windSpeed":"wind speed","waveHeight":"wave height",
                     "currentMag":"current magnitude","rainfall":"rainfall"}[var]
            contribs.append({"variable": var, "value": jround(val), "points": pts,
                             "text": f"Elevated {label} contributed {pts:g} points"})
    if available == 0:
        return {"score": None, "category": "Insufficient Data", "contributions": [], "inputsUsed": 0}
    score = round(min(total, 100), 1)
    cat = "Low Risk" if score < 34 else "Moderate Risk" if score < 67 else "High Risk"
    contribs.sort(key=lambda c: -c["points"])
    return {"score": score, "category": cat, "contributions": contribs, "inputsUsed": available}

# Per-hour risk series for the lake stations (attach nearest verified wind + rainfall = Bteha)
bteha_h = hourly(meteo["bteha"]["df"], ["windSpeed","rainfall"]).rename(
    columns={"windSpeed":"windSpeed","rainfall":"rainfall"})

def risk_series(lake_df, has_waves):
    d = lake_df.merge(bteha_h, on="timestamp", how="left")
    recs = []
    for _, r in d.iterrows():
        inp = {"windSpeed": r.get("windSpeed"), "currentMag": r.get("currentMag"),
               "rainfall": r.get("rainfall")}
        if has_waves: inp["waveHeight"] = r.get("waveHeight")
        rk = compute_risk(inp)
        recs.append({"timestamp": iso(r["timestamp"]), **rk})
    return recs

knw_risk = risk_series(knw_lake, True)
knc_risk = risk_series(knc_lake, False)

# ===========================================================================
# STATIONS.JSON  (status, latest obs, latest risk, variables, missing %)
# ===========================================================================
def meteo_missing_pct(m):
    df = m["df"]; cols = m["verified"]
    if df.empty or not cols: return 100.0
    cells = df[cols]
    return jround(100.0 * cells.isna().sum().sum() / (cells.shape[0]*cells.shape[1]), 2)

def latest_nonnull(df, col):
    if col not in df or df[col].dropna().empty: return None, None
    sub = df.dropna(subset=[col])
    last = sub.iloc[-1]
    return jround(last[col]), iso(last["timestamp"])

stations_out = []
DATASET_MIN, DATASET_MAX = [], []

for sid, meta in STATIONS.items():
    entry = {"id": sid, "name": meta["name"], "type": meta["type"],
             "lat": meta["lat"], "lng": meta["lng"], "approxCoords": meta["approxCoords"],
             "assumptions": [], "unverified": []}
    if sid in meteo:
        m = meteo[sid]; df = m["df"]
        variables = []
        for v in ["windSpeed","windDir","airTemp","humidity","rainfall","waterTemp","lightLevel","pressure",
                  "spCond","dissolvedO2","turbidity","chlorophyll","orp"]:
            if v in m["verified"] and v in df and df[v].notna().any():
                variables.append(v)
        latest = {}
        for v in variables:
            val, t = latest_nonnull(df, v)
            latest[v] = {"value": val, "unit": UNITS[v], "at": t}
        rng = [iso(df["timestamp"].min()), iso(df["timestamp"].max())]
        entry.update({"records": int(len(df)), "dateRange": rng,
                      "variables": variables, "missingPct": meteo_missing_pct(m),
                      "lastTimestamp": rng[1], "latest": latest})
        if m["ws_assumed"]:
            entry["assumptions"].append("Wind speed unit assumed m/s (undocumented in source).")
        if m.get("daily_summary"):
            entry["assumptions"].append("Daily summary data (not hourly); timestamps set to noon of each day.")
        for u in m["unverified"]:
            entry["unverified"].append(u)
        # status
        miss = entry["missingPct"]
        entry["status"] = "Available" if miss < 5 else "Warning" if miss < 40 else "Missing"
        # partial risk for meteo (wind + rainfall only)
        winp, _ = latest_nonnull(df, "windSpeed")
        rnp, _ = latest_nonnull(df, "rainfall")
        rk = compute_risk({"windSpeed": winp, "rainfall": rnp}) if "windSpeed" in variables else \
             {"score": None, "category": "Insufficient Data", "contributions": [], "inputsUsed": 0}
        entry["risk"] = rk
        if sid == "a_probe":
            entry["dataQualityNote"] = ("Water-quality probe. Measures dissolved oxygen, chlorophyll, "
                                        "turbidity, specific conductance, ORP and water temperature. "
                                        "Risk model not applicable (no wind or current data).")
        else:
            entry["dataQualityNote"] = ("Meteorological station. Risk uses verified wind/rainfall only; "
                                        "no wave or current data at this site.")
        DATASET_MIN.append(df["timestamp"].min()); DATASET_MAX.append(df["timestamp"].max())
    elif sid == "knw":
        df = knw_lake
        vis = ["waveHeight","waveHeightMax","waveHeightTop10","wavePeriod","wavePeriodMean",
               "waveDir","currentMag","currentDir","sensorDepth"]
        variables = [v for v in vis if v in df and df[v].notna().any()]
        latest = {}
        for v in variables:
            val, t = latest_nonnull(df, v); latest[v] = {"value": val, "unit": UNITS[v], "at": t}
        rng = [iso(df["timestamp"].min()), iso(df["timestamp"].max())]
        miss = jround(100.0*df[variables].isna().sum().sum()/(df.shape[0]*len(variables)),2)
        entry.update({"records": int(len(df)), "dateRange": rng, "variables": variables,
                      "missingPct": miss, "lastTimestamp": rng[1], "latest": latest,
                      "status": "Available" if miss < 30 else "Warning",
                      "risk": knw_risk[-1] if knw_risk else {"score":None,"category":"Insufficient Data","contributions":[]},
                      "dataQualityNote": "Golan Beach: significant wave height + current profile (surface used as default). "
                                         "Depth(mm) is a sensor field, not lake level. Wind/rainfall borrowed from Bteha for risk."})
        DATASET_MIN.append(df["timestamp"].min()); DATASET_MAX.append(df["timestamp"].max())
    elif sid == "knc":
        df = knc_lake; vis = ["currentMag","currentDir"]
        variables = [v for v in vis if v in df and df[v].notna().any()]
        latest = {}
        for v in variables:
            val, t = latest_nonnull(df, v); latest[v] = {"value": val, "unit": UNITS[v], "at": t}
        rng = [iso(df["timestamp"].min()), iso(df["timestamp"].max())]
        miss = jround(100.0*df[variables].isna().sum().sum()/(df.shape[0]*len(variables)),2)
        entry.update({"records": int(len(df)), "dateRange": rng, "variables": variables,
                      "missingPct": miss, "lastTimestamp": rng[1], "latest": latest,
                      "status": "Available" if miss < 30 else "Warning",
                      "risk": knc_risk[-1] if knc_risk else {"score":None,"category":"Insufficient Data","contributions":[]},
                      "dataQualityNote": "Station F: current profile only (surface used as default). "
                                         "No wave data; wind/rainfall borrowed from Bteha for risk."})
        DATASET_MIN.append(df["timestamp"].min()); DATASET_MAX.append(df["timestamp"].max())
    stations_out.append(entry)

dataset_range = [iso(min(DATASET_MIN)), iso(max(DATASET_MAX))]
latest_obs = max(iso(x) for x in DATASET_MAX)

# ===========================================================================
# TIMESERIES.JSON  (daily + hourly aggregates per station)
# ===========================================================================
ts_out = {}
ALL_METEO_VARS = ["windSpeed","airTemp","humidity","rainfall","waterTemp","lightLevel","pressure",
                  "spCond","dissolvedO2","turbidity","chlorophyll","orp"]
for sid in meteo:
    df = meteo[sid]["df"]
    cols = [c for c in ALL_METEO_VARS if c in df and df[c].notna().any()]
    ts_out[sid] = {"daily": aggregate(df, cols, "1D"), "hourly": aggregate(df, cols, "1h"), "vars": cols}
_knw_ts_vars = ["waveHeight","currentMag"] + [
    c for c in ["waveHeightMax","waveHeightTop10","wavePeriodMean"]
    if c in knw_lake.columns and knw_lake[c].notna().any()
]
ts_out["knw"] = {"daily": aggregate(knw_lake, _knw_ts_vars, "1D"),
                 "hourly": aggregate(knw_lake, _knw_ts_vars, "1h"),
                 "vars": _knw_ts_vars}
ts_out["knc"] = {"daily": aggregate(knc_lake, ["currentMag"], "1D"),
                 "hourly": aggregate(knc_lake, ["currentMag"], "1h"), "vars": ["currentMag"]}

# ===========================================================================
# ALERTS.JSON  (from risk series; high/moderate events)
# ===========================================================================
def build_alerts(series, station_name, station_id):
    alerts = []
    for r in series:
        if r["category"] in ("Moderate Risk","High Risk") and r["contributions"]:
            top = r["contributions"][0]
            var = top["variable"]; t = THRESH.get(var, {})
            alerts.append({
                "id": f"{station_id}-{r['timestamp']}",
                "station": station_name, "stationId": station_id,
                "timestamp": r["timestamp"], "severity": r["category"],
                "variable": var, "value": top["value"],
                "threshold": (t or {}).get("p85"),
                "score": r["score"],
                "explanation": "; ".join(c["text"] for c in r["contributions"]),
                "recommendation": "Elevated environmental risk detected. This is a research prototype; "
                                  "do not rely on it for real-world safety decisions. Consult official "
                                  "authorities and lifeguard instructions."})
    return alerts

all_alerts = build_alerts(knw_risk, STATIONS["knw"]["name"], "knw") + \
             build_alerts(knc_risk, STATIONS["knc"]["name"], "knc")
all_alerts.sort(key=lambda a: a["timestamp"], reverse=True)
alerts_summary = {
    "total": len(all_alerts),
    "high": sum(1 for a in all_alerts if a["severity"] == "High Risk"),
    "moderate": sum(1 for a in all_alerts if a["severity"] == "Moderate Risk"),
}

# ===========================================================================
# STATISTICS.JSON
# ===========================================================================
ALPHA = 0.05
def interp_p(p): return "statistically significant (reject H0)" if p < ALPHA else "not statistically significant (fail to reject H0)"

def align(df_a, col_a, df_b, col_b):
    a = df_a[["timestamp", col_a]].dropna()
    b = df_b[["timestamp", col_b]].dropna()
    m = pd.merge(a, b, on="timestamp", how="inner").dropna()
    return m[col_a].values, m[col_b].values, len(m)

statistics = {"alpha": ALPHA, "tests": [], "descriptive": {}, "correlationMatrix": None}

# Descriptive stats for verified variables
desc = {}
desc["Bteha wind speed (m/s)"] = describe(meteo["bteha"]["df"]["windSpeed"])
desc["Zemah wind speed (m/s)"] = describe(meteo["zemah"]["df"]["windSpeed"])
desc["Ein Gev wind speed (m/s)"] = describe(meteo["eingev"]["df"]["windSpeed"])
desc["Bteha air temp (C)"] = describe(meteo["bteha"]["df"]["airTemp"])
desc["Ginosar air temp (C)"] = describe(meteo["ginosar"]["df"]["airTemp"])
desc["Metr-A water temp (C)"] = describe(meteo["metr_a"]["df"]["waterTemp"])
desc["Metr-A wind speed (m/s)"] = describe(meteo["metr_a"]["df"]["windSpeed"])
desc["A-Probe water temp (C)"] = describe(meteo["a_probe"]["df"]["waterTemp"])
desc["A-Probe dissolved oxygen (mg/L)"] = describe(meteo["a_probe"]["df"]["dissolvedO2"])
desc["A-Probe chlorophyll (µg/L)"] = describe(meteo["a_probe"]["df"]["chlorophyll"])
desc["A-Probe turbidity (FNU)"] = describe(meteo["a_probe"]["df"]["turbidity"])
desc["KNW wave height Hs (m)"] = describe(knw_wav["df"]["waveHeight"])
desc["KNW current magnitude (cm/s)"] = describe(knw_cur["df"]["currentMag"])
desc["KNC current magnitude (cm/s)"] = describe(knc_cur["df"]["currentMag"])
statistics["descriptive"] = desc

# A. Pearson: wind (Bteha) vs Hs (KNW)
def scatter_sample(x, y, n=400):
    if len(x) <= n: idx = np.arange(len(x))
    else: idx = np.linspace(0, len(x)-1, n).astype(int)
    return [[jround(float(x[i])), jround(float(y[i]))] for i in idx]

def regression_line(x, y):
    if len(x) < 2: return None
    sl, ic = np.polyfit(x, y, 1)
    xs = [float(np.min(x)), float(np.max(x))]
    return {"slope": jround(sl), "intercept": jround(ic),
            "points": [[jround(xs[0]), jround(sl*xs[0]+ic)], [jround(xs[1]), jround(sl*xs[1]+ic)]]}

x, y, n = align(bteha_h.rename(columns={"windSpeed":"windSpeed"}).assign(timestamp=bteha_h["timestamp"]),
                "windSpeed",
                hourly(knw_wav["df"], ["waveHeight"]), "waveHeight")
if n > 2:
    r, p = sps.pearsonr(x, y)
    statistics["tests"].append({
        "id":"pearson_wind_wave","name":"Pearson correlation",
        "title":"Wind speed (Bteha) vs significant wave height (Golan Beach)",
        "why":"Both variables are continuous and we expect an approximately linear physical relationship (wind forcing drives wave growth).",
        "variables":[{"name":"Wind speed (Bteha)","type":"continuous","unit":"m/s"},
                     {"name":"Significant wave height Hs (KNW)","type":"continuous","unit":"m"}],
        "assumptions":["Linear relationship","Approx. bivariate normality","Independent paired hourly observations"],
        "h0":"There is no linear correlation between wind speed and wave height (rho = 0).",
        "h1":"There is a non-zero linear correlation (rho != 0).",
        "n":n,"statistic":jround(r),"statisticName":"r","pValue":jround(p,6),
        "interpretation":interp_p(p),
        "plainLanguage":f"Hourly wind speed and wave height show an r of {r:.2f} over {n} aligned hours.",
        "scatter":scatter_sample(x,y),"regression":regression_line(x,y),
        "xLabel":"Wind speed (m/s)","yLabel":"Hs (m)"})

# wind vs current magnitude (Bteha vs KNW surface)
x2, y2, n2 = align(bteha_h, "windSpeed", hourly(knw_cur["df"], ["currentMag"]), "currentMag")
if n2 > 2:
    r2, p2 = sps.pearsonr(x2, y2)
    statistics["tests"].append({
        "id":"pearson_wind_current","name":"Pearson correlation",
        "title":"Wind speed (Bteha) vs surface current magnitude (Golan Beach)",
        "why":"Both continuous; testing whether wind forcing is linearly associated with near-surface currents.",
        "variables":[{"name":"Wind speed (Bteha)","type":"continuous","unit":"m/s"},
                     {"name":"Surface current magnitude (KNW)","type":"continuous","unit":"cm/s"}],
        "assumptions":["Linear relationship","Approx. bivariate normality","Independent paired hourly observations"],
        "h0":"No linear correlation (rho = 0).","h1":"Non-zero linear correlation (rho != 0).",
        "n":n2,"statistic":jround(r2),"statisticName":"r","pValue":jround(p2,6),
        "interpretation":interp_p(p2),
        "plainLanguage":f"Hourly wind and surface current show an r of {r2:.2f} over {n2} aligned hours.",
        "scatter":scatter_sample(x2,y2),"regression":regression_line(x2,y2),
        "xLabel":"Wind speed (m/s)","yLabel":"Current (cm/s)"})

# B. Spearman: wind vs Hs (monotonic / non-normal robust)
if n > 2:
    rs, ps = sps.spearmanr(x, y)
    statistics["tests"].append({
        "id":"spearman_wind_wave","name":"Spearman rank correlation",
        "title":"Wind speed vs wave height (rank-based)",
        "why":"Wind and wave-height distributions are right-skewed and the relationship may be monotonic rather than strictly linear; Spearman is robust to both.",
        "variables":[{"name":"Wind speed (Bteha)","type":"continuous"},
                     {"name":"Hs (KNW)","type":"continuous"}],
        "assumptions":["Monotonic relationship","Ordinal/continuous paired data"],
        "h0":"No monotonic association (rho_s = 0).","h1":"A monotonic association exists (rho_s != 0).",
        "n":n,"statistic":jround(rs),"statisticName":"rho_s","pValue":jround(ps,6),
        "interpretation":interp_p(ps),
        "plainLanguage":f"Rank correlation rho_s = {rs:.2f} over {n} aligned hours."})

# C. One-way ANOVA: wind speed across verified meteo stations (+ Kruskal fallback)
groups = {"Bteha": meteo["bteha"]["df"]["windSpeed"].dropna().values,
          "Zemah": meteo["zemah"]["df"]["windSpeed"].dropna().values,
          "Ein Gev": meteo["eingev"]["df"]["windSpeed"].dropna().values}
gvals = list(groups.values())
F, pA = sps.f_oneway(*gvals)
# Levene for equal-variance assumption
W, pL = sps.levene(*gvals)
norm_ok = all(len(v) > 5000 for v in gvals)  # large-n CLT note
H, pK = sps.kruskal(*gvals)
statistics["tests"].append({
    "id":"anova_wind_station","name":"One-way ANOVA",
    "title":"Mean wind speed by meteorological station",
    "why":"Comparing the mean of one continuous variable (wind speed) across three independent groups (stations).",
    "variables":[{"name":"Station","type":"categorical (3 groups)"},
                 {"name":"Wind speed","type":"continuous","unit":"m/s"}],
    "assumptions":["Independent groups","Approx. normal residuals (large n -> CLT)","Homogeneity of variance (Levene)"],
    "assumptionChecks":{"leveneStat":jround(W),"leveneP":jround(pL,6),
                        "homogeneity": "violated (p<0.05)" if pL<ALPHA else "ok"},
    "h0":"All station mean wind speeds are equal (mu1 = mu2 = mu3).",
    "h1":"At least one station mean differs.",
    "n":int(sum(len(v) for v in gvals)),"statistic":jround(F),"statisticName":"F","pValue":jround(pA,6),
    "groupMeans":{k: jround(np.mean(v)) for k,v in groups.items()},
    "interpretation":interp_p(pA),
    "nonParametric":{"name":"Kruskal-Wallis (fallback)","statistic":jround(H),"H":jround(H),
                     "pValue":jround(pK,6),"interpretation":interp_p(pK),
                     "shown": pL < ALPHA},
    "plainLanguage":f"F = {F:.1f}, p = {pA:.3g}. Levene p = {pL:.3g} -> variance homogeneity "
                    f"{'violated, Kruskal-Wallis reported as well' if pL<ALPHA else 'acceptable'}."})

# D. Chi-square: month vs risk category (KNW lake risk)
risk_df = pd.DataFrame(knw_risk)
risk_df = risk_df[risk_df["category"] != "Insufficient Data"].copy()
risk_df["month"] = pd.to_datetime(risk_df["timestamp"]).dt.strftime("%Y-%m")
ct = pd.crosstab(risk_df["month"], risk_df["category"])
order = [c for c in ["Low Risk","Moderate Risk","High Risk"] if c in ct.columns]
ct = ct[order]
chi2, pC, dof, expected = sps.chi2_contingency(ct.values)
low_exp = int((expected < 5).sum())
statistics["tests"].append({
    "id":"chi2_month_risk","name":"Chi-square test of independence",
    "title":"Month vs calculated risk category (Golan Beach)",
    "why":"Both variables are categorical (month, ordinal risk class); chi-square tests whether risk-class frequencies depend on month.",
    "variables":[{"name":"Month","type":"categorical"},
                 {"name":"Calculated risk category","type":"categorical (ordinal)"}],
    "assumptions":["Independent observations","Expected count >= 5 in most cells"],
    "h0":"Risk category is independent of month.","h1":"Risk category depends on month.",
    "n":int(ct.values.sum()),"statistic":jround(chi2),"statisticName":"chi2","pValue":jround(pC,6),
    "dof":int(dof),"expectedLowCells":low_exp,
    "expectedWarning": low_exp>0,
    "contingency":{"rows":list(ct.index),"cols":order,"values":ct.values.tolist()},
    "interpretation":interp_p(pC),
    "plainLanguage":f"chi2 = {chi2:.1f}, dof = {dof}, p = {pC:.3g}."
                    + (f" Warning: {low_exp} expected cells < 5." if low_exp else "")})

# E. Linear regression: predict Hs from wind (train/test split)
if n > 30:
    from sklearn.model_selection import train_test_split as _tts
    X = x.reshape(-1,1); Y = y
    Xtr,Xte,Ytr,Yte = _tts(X, Y, test_size=0.3, random_state=42, shuffle=True)
    lr = LinearRegression().fit(Xtr,Ytr)
    pred = lr.predict(Xte)
    resid = (Yte - pred)
    av = [[jround(float(Yte[i])), jround(float(pred[i]))] for i in np.linspace(0,len(Yte)-1,min(300,len(Yte))).astype(int)]
    rp = [[jround(float(pred[i])), jround(float(resid[i]))] for i in np.linspace(0,len(pred)-1,min(300,len(pred))).astype(int)]
    statistics["tests"].append({
        "id":"linreg_hs_wind","name":"Linear regression",
        "title":"Predict significant wave height from wind speed",
        "why":"Interpretable model quantifying how Hs changes per unit wind speed; continuous predictor and response.",
        "variables":[{"name":"Wind speed (Bteha)","type":"predictor (continuous)","unit":"m/s"},
                     {"name":"Hs (KNW)","type":"response (continuous)","unit":"m"}],
        "assumptions":["Linearity","Independent residuals","Homoscedasticity","Approx. normal residuals"],
        "equation":f"Hs = {lr.coef_[0]:.4f} * wind + {lr.intercept_:.4f}",
        "coefficients":{"wind": jround(lr.coef_[0],4),"intercept": jround(lr.intercept_,4)},
        "r2": jround(r2_score(Yte,pred)),"mae": jround(mean_absolute_error(Yte,pred)),
        "rmse": jround(math.sqrt(mean_squared_error(Yte,pred))),
        "trainPeriod": n,"trainN": int(len(Xtr)),"testN": int(len(Xte)),
        "actualVsPredicted": av,"residuals": rp,
        "interpretation":"Caution: the calculated risk score is NOT used here as ground truth; "
                         "this model uses measured Hs only.",
        "plainLanguage":f"Each +1 m/s of wind is associated with ~{lr.coef_[0]:.3f} m more Hs; "
                        f"test R^2 = {r2_score(Yte,pred):.2f}."})

# Correlation matrix (verified, hourly aligned where possible)
cm_vars = {"Wind (Bteha)": bteha_h.set_index("timestamp")["windSpeed"],
           "Hs (KNW)": hourly(knw_wav["df"], ["waveHeight"]).set_index("timestamp")["waveHeight"],
           "Current (KNW)": hourly(knw_cur["df"], ["currentMag"]).set_index("timestamp")["currentMag"],
           "Rain (Bteha)": bteha_h.set_index("timestamp")["rainfall"]}
cm_df = pd.DataFrame(cm_vars)
corr = cm_df.corr(method="pearson")
statistics["correlationMatrix"] = {"labels": list(corr.columns),
                                    "values": [[jround(v) for v in row] for row in corr.values.tolist()]}

# ===========================================================================
# RANDOM FOREST (ML extension): predict Hs from met features
# ===========================================================================
ml = None
rf_df = bteha_h.merge(hourly(knw_wav["df"], ["waveHeight"]), on="timestamp", how="inner")
rf_df = rf_df.merge(hourly(meteo["bteha"]["df"], ["windDir","humidity"]), on="timestamp", how="left")
rf_df["hour"] = pd.to_datetime(rf_df["timestamp"]).dt.hour
rf_df["month"] = pd.to_datetime(rf_df["timestamp"]).dt.month
rf_df["wdir_sin"] = np.sin(np.deg2rad(rf_df["windDir"]))
rf_df["wdir_cos"] = np.cos(np.deg2rad(rf_df["windDir"]))
feat = ["windSpeed","wdir_sin","wdir_cos","humidity","hour","month"]
rf_df = rf_df.dropna(subset=feat+["waveHeight"])
if len(rf_df) > 50:
    from sklearn.model_selection import train_test_split
    Xall, Yall = rf_df[feat], rf_df["waveHeight"]
    Xtr,Xte,Ytr,Yte = train_test_split(Xall, Yall, test_size=0.3, random_state=42, shuffle=True)
    rf = RandomForestRegressor(n_estimators=200, random_state=42, n_jobs=-1).fit(Xtr,Ytr)
    pred = rf.predict(Xte)
    Yte = Yte.reset_index(drop=True)
    av = [[jround(float(Yte.iloc[i])), jround(float(pred[i]))] for i in np.linspace(0,len(Yte)-1,min(300,len(Yte))).astype(int)]
    ml = {"target":"Significant wave height Hs (m)","model":"RandomForestRegressor (200 trees)",
          "features": feat, "trainN": int(len(Xtr)), "testN": int(len(Xte)),
          "split":"random 70/30 shuffle (random_state=42)",
          "trainPeriod":[iso(rf_df["timestamp"].iloc[0]), iso(rf_df["timestamp"].iloc[-1])],
          "testPeriod":"30% random hold-out across the full period",
          "mae": jround(mean_absolute_error(Yte,pred)),
          "rmse": jround(math.sqrt(mean_squared_error(Yte,pred))),
          "r2": jround(r2_score(Yte,pred)),
          "featureImportance": sorted([{"feature":f,"importance":jround(im)} for f,im in zip(feat, rf.feature_importances_)],
                                      key=lambda d:-d["importance"]),
          "actualVsPredicted": av,
          "note":"Regression of a measured physical quantity (Hs). No safety-risk classifier is trained; "
                 "the rule-based risk score is never used as a training label. Random hold-out is used for "
                 "feature-importance demonstration; note hourly observations are temporally autocorrelated, "
                 "so the hold-out R^2 is optimistic relative to forecasting truly unseen periods."}

# ===========================================================================
# DATA QUALITY.JSON
# ===========================================================================
def dup_and_gaps(df, expected_freq_min):
    if df.empty: return 0, None
    dups = int(df["timestamp"].duplicated().sum())
    return dups, None

dq_files = []
def dq_meteo(sid, m, expected_min):
    df = m["df"]
    dups = int(df["timestamp"].duplicated().sum())
    missing = {}
    verified_present = [v for v in m["verified"] if v in df]
    for v in verified_present:
        missing[v] = int(df[v].isna().sum())
    # timestamp gaps
    diffs = df["timestamp"].sort_values().diff().dropna()
    gaps = int((diffs > pd.Timedelta(minutes=expected_min*3)).sum())
    usable = jround(100.0*(1 - df[verified_present].isna().mean().mean())) if verified_present else 0
    return {"file": m["filename"], "station": STATIONS[sid]["name"], "parsed": True,
            "rows": int(len(df)), "rawRows": m["raw_rows"],
            "dateRange":[iso(df["timestamp"].min()), iso(df["timestamp"].max())],
            "missingPerColumn": missing, "duplicateTimestamps": dups,
            "timestampGaps": gaps, "usablePct": usable,
            "unverifiedColumns": m["unverified"],
            "depthLevels": None}

for sid in meteo:
    expected_min = 1440 if meteo[sid].get("daily_summary") else 10
    dq_files.append(dq_meteo(sid, meteo[sid], expected_min))

# lake currents quality (per concatenated station)
for sid, cur, expected_levels in [("knw", knw_cur, None), ("knc", knc_cur, None)]:
    df = cur["df"]
    dq_files.append({"file": f"{sid.upper()}*_Currents.csv ({len(cur['per_file'])} deployments)",
                     "station": STATIONS[sid]["name"], "parsed": True,
                     "rows": int(len(df)), "rawRows": cur["raw_rows"],
                     "dateRange":[iso(df["timestamp"].min()), iso(df["timestamp"].max())],
                     "missingPerColumn":{"currentMag": int(df["currentMag"].isna().sum())},
                     "duplicateTimestamps": int(df["timestamp"].duplicated().sum()),
                     "timestampGaps": None,
                     "usablePct": jround(100.0*(1-df["currentMag"].isna().mean())),
                     "depthLevels": cur["depth_levels"][:60],
                     "unverifiedColumns": [],
                     "deployments": cur["per_file"]})
# KNW waves quality
wdf = knw_wav["df"]
dq_files.append({"file": f"KNW*_Waves.xlsx ({len(knw_wav['per_file'])} deployments)",
                 "station": STATIONS["knw"]["name"], "parsed": True,
                 "rows": int(len(wdf)), "rawRows": knw_wav["raw_rows"],
                 "dateRange":[iso(wdf["timestamp"].min()), iso(wdf["timestamp"].max())],
                 "missingPerColumn":{"waveHeight": int(wdf["waveHeight"].isna().sum())},
                 "duplicateTimestamps": int(wdf["timestamp"].duplicated().sum()),
                 "timestampGaps": None,
                 "usablePct": jround(100.0*(1-wdf["waveHeight"].isna().mean())),
                 "depthLevels": None, "unverifiedColumns": [],
                 "deployments": knw_wav["per_file"],
                 "note":"sensorDepth column (Depth(mm), ~18.8-20.0 m) is a sensor measurement field, NOT lake level."})

data_quality = {
    "filesParsed": sum(1 for f in dq_files if f["parsed"]),
    "filesFailed": sum(1 for f in dq_files if not f["parsed"]),
    "files": dq_files,
    "unverifiedColumns": EXCLUSIONS,
    "unconfirmedUnits":[
        {"field":"Bteha/Zemah wind speed","status":"assumed m/s","detail":"Source files do not document the unit; values match the explicitly m/s stations (Ein Gev, Ginosar)."},
        {"field":"Ein Gev air temperature","status":"unverified","detail":"Impossible as Celsius; possibly Fahrenheit or miscalibrated. Excluded from analysis."},
        {"field":"Ginosar wind speed","status":"unverified","detail":"Stuck near 0 m/s all year; likely failed anemometer. Excluded from analysis."},
        {"field":"Wave Depth(mm)","status":"sensor field","detail":"~18.8-20.0 m water-column/instrument depth; NOT Lake Kinneret water level."},
        {"field":"Timezone (meteo/currents)","status":"assumed local","detail":"Unlabeled; assumed Israel local. Waves are GMT and were converted to local."},
    ],
    "parseErrors": parse_errors,
}

# ===========================================================================
# MANIFEST + WRITE
# ===========================================================================
manifest = {
    "generatedAt": dt.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
    "datasetRange": dataset_range,
    "latestObservation": latest_obs,
    "riskVersion": RISK_VERSION,
    "stationCount": len(stations_out),
    "assumptions":[
        "Bteha & Zemah wind speed treated as m/s (undocumented in source).",
        "Wave timestamps converted from GMT to Israel local time to align with meteo/current series.",
        "Meteo & current timestamps assumed Israel local (unlabeled in source).",
        "Current 'magnitude' per station uses the shallowest (surface) depth as default; a depth-averaged value is also computed.",
        "Map marker positions are approximate, for the stylized map only -- not survey GPS.",
        "Metr-A (2025): daily summary data; timestamps assigned to noon. Light level unit unconfirmed (displayed as-is).",
        "A-Probe (2025): daily water-quality summary (DO, chlorophyll, turbidity, SpCond, ORP); timestamps assigned to noon.",
        "KNW08/KNW09 (2025): daily wave summaries extend the hourly KNW wave record. Mar-12 KNW09 deployment artifact excluded (KNW08 used for that day).",
    ],
    "excludedColumns":[f"{e['station']}.{e['column']} ({e['variable']}) -- {e['action']}" for e in EXCLUSIONS],
    "disclaimer":"This system is an academic research prototype. Its risk classifications are not official "
                 "swimming or emergency-safety instructions.",
}

def _json_default(o):
    if isinstance(o, (np.integer,)): return int(o)
    if isinstance(o, (np.floating,)):
        return None if (math.isnan(o) or math.isinf(o)) else float(o)
    if isinstance(o, (np.bool_,)): return bool(o)
    if isinstance(o, (np.ndarray,)): return o.tolist()
    raise TypeError(f"not serializable: {type(o)}")

def write(name, obj):
    p = OUT / name
    p.write_text(json.dumps(obj, ensure_ascii=False, separators=(",",":"), default=_json_default))
    print(f"  wrote {name}  ({p.stat().st_size/1024:.0f} KB)")

print("Writing processed JSON ...")
write("manifest.json", manifest)
write("stations.json", stations_out)
write("timeseries.json", ts_out)
write("dataQuality.json", data_quality)
write("alerts.json", {"summary": alerts_summary, "alerts": all_alerts[:500],
                       "thresholds": THRESH, "riskVersion": RISK_VERSION})
write("risk.json", {"version": RISK_VERSION, "weights": RISK_WEIGHTS, "thresholds": THRESH,
                    "method":"Data-driven percentile thresholds (p50 / p85) on verified variables; "
                             "points scale 0 below p50, up to half-weight by p85, up to full weight above.",
                    "categories":{"Low Risk":"score < 34","Moderate Risk":"34-66","High Risk":">= 67",
                                  "Insufficient Data":"no verified inputs available"},
                    "knwSeries": knw_risk[-2000:], "kncSeries": knc_risk[-2000:]})
write("statistics.json", {**statistics, "ml": ml})
print("DONE.")
